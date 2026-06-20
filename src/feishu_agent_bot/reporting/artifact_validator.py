from __future__ import annotations

import hashlib
import csv
import json
import math
import re
import zipfile
from dataclasses import asdict, is_dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..errors import ArtifactValidationError as _ArtifactValidationError


class ArtifactValidationError(_ArtifactValidationError, ValueError):
    pass


class ArtifactValidator:
    def validate_report_ir(self, ir) -> dict[str, object]:
        if is_dataclass(ir):
            payload = asdict(ir)
        elif isinstance(ir, dict):
            payload = ir
        else:
            raise ArtifactValidationError("Report IR must be an object")
        self._validate_report_ir_schema(payload)
        self.validate_no_sensitive_text(
            json.dumps(payload, ensure_ascii=False), location="Report IR"
        )
        return payload

    def validate_report_ir_json(self, path: str | Path) -> str:
        path = Path(path)
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise ArtifactValidationError(
                f"Report IR JSON cannot be read: {exc}"
            ) from exc
        self.validate_report_ir(payload)
        return self.sha256(path)

    def _validate_report_ir_schema(self, payload: dict[str, object]) -> None:
        for field_name in ("job_id", "title", "generated_at", "purpose"):
            value = payload.get(field_name)
            if not isinstance(value, str) or not value.strip():
                raise ArtifactValidationError(
                    f"Report IR field must be a non-empty string: {field_name}"
                )
        version = payload.get("version")
        if not isinstance(version, int) or isinstance(version, bool) or version < 1:
            raise ArtifactValidationError(
                "Report IR version must be a positive integer"
            )
        if not isinstance(payload.get("methodology"), dict):
            raise ArtifactValidationError("Report IR methodology must be an object")
        if not isinstance(payload.get("metadata"), dict):
            raise ArtifactValidationError("Report IR metadata must be an object")

        list_fields = (
            "executive_summary",
            "sections",
            "tables",
            "charts",
            "analysis_results",
            "change_events",
            "claims",
            "evidence_references",
            "data_quality",
            "limitations",
            "sources",
            "appendices",
        )
        for field_name in list_fields:
            if not isinstance(payload.get(field_name), list):
                raise ArtifactValidationError(
                    f"Report IR field must be an array: {field_name}"
                )

        sections = self._mapping_items(payload, "sections")
        appendices = self._mapping_items(payload, "appendices")
        tables = self._mapping_items(payload, "tables")
        charts = self._mapping_items(payload, "charts")
        analysis_results = self._mapping_items(payload, "analysis_results")
        claims = self._mapping_items(payload, "claims")
        evidence = self._mapping_items(payload, "evidence_references")
        executive_summary = self._mapping_items(payload, "executive_summary")
        change_events = self._mapping_items(payload, "change_events")

        for index, section in enumerate(sections + appendices):
            if (
                not isinstance(section.get("title"), str)
                or not section["title"].strip()
            ):
                raise ArtifactValidationError(
                    f"Report IR section at index {index} has no title"
                )
            if not isinstance(section.get("body"), str):
                raise ArtifactValidationError(
                    f"Report IR section at index {index} body must be a string"
                )
            if not section["body"].strip():
                raise ArtifactValidationError(
                    f"Report IR section at index {index} body must not be empty"
                )
        self._require_reference_arrays(
            sections + appendices,
            "claim_ids",
            "evidence_ids",
            "analysis_result_ids",
            "table_ids",
            "chart_ids",
        )
        self._require_reference_arrays(
            claims,
            "supporting_evidence_ids",
            "contradicting_evidence_ids",
        )
        self._require_reference_arrays(
            analysis_results, "input_evidence_ids", "input_dataset_ids"
        )
        self._require_reference_arrays(charts, "analysis_result_ids")

        self._unique_ids(sections + appendices, "section_id", "section")
        table_ids = self._unique_ids(tables, "table_id", "table")
        chart_ids = self._unique_ids(charts, "chart_id", "chart")
        analysis_ids = self._unique_ids(
            analysis_results, ("analysis_result_id", "result_id"), "analysis result"
        )
        claim_ids = self._unique_ids(claims, "claim_id", "claim")
        evidence_ids = self._unique_ids(evidence, "evidence_id", "evidence")

        section_like = sections + appendices + executive_summary
        self._assert_references(
            "claim", self._references(section_like, "claim_ids", "claim_id"), claim_ids
        )
        self._assert_references(
            "evidence",
            self._references(section_like, "evidence_ids", "evidence_id"),
            evidence_ids,
        )
        self._assert_references(
            "analysis result",
            self._references(
                section_like, "analysis_result_ids", "analysis_result_id"
            ),
            analysis_ids,
        )
        self._assert_references(
            "table", self._references(section_like, "table_ids", "table_id"), table_ids
        )
        self._assert_references(
            "chart", self._references(section_like, "chart_ids", "chart_id"), chart_ids
        )

        evidence_referrers = claims + analysis_results + change_events
        evidence_refs = self._references(
            evidence_referrers,
            "supporting_evidence_ids",
            "contradicting_evidence_ids",
            "input_evidence_ids",
            "evidence_ids",
        )
        self._assert_references("evidence", evidence_refs, evidence_ids)
        table_analysis_refs = self._references(
            tables, "analysis_result_ids", "analysis_result_id"
        )
        chart_analysis_refs = self._references(
            charts, "analysis_result_ids", "analysis_result_id"
        )
        self._assert_references(
            "analysis result",
            table_analysis_refs + chart_analysis_refs,
            analysis_ids,
        )

    @staticmethod
    def _mapping_items(
        payload: dict[str, object], field_name: str
    ) -> list[dict[str, object]]:
        values = payload[field_name]
        for index, item in enumerate(values):
            if not isinstance(item, dict):
                raise ArtifactValidationError(
                    f"Report IR {field_name}[{index}] must be an object"
                )
        return values

    @staticmethod
    def _unique_ids(
        items: list[dict[str, object]],
        field_names: str | tuple[str, ...],
        label: str,
    ) -> set[str]:
        names = (field_names,) if isinstance(field_names, str) else field_names
        values = []
        for index, item in enumerate(items):
            value = next((item.get(name) for name in names if item.get(name)), None)
            if value is None or not str(value).strip():
                raise ArtifactValidationError(
                    f"Report IR {label} at index {index} has no ID"
                )
            values.append(str(value))
        duplicates = sorted({value for value in values if values.count(value) > 1})
        if duplicates:
            raise ArtifactValidationError(
                f"Report IR has duplicate {label} IDs: {', '.join(duplicates)}"
            )
        return set(values)

    @staticmethod
    def _references(
        items: list[dict[str, object]], *field_names: str
    ) -> list[str]:
        references = []
        for item in items:
            for field_name in field_names:
                value = item.get(field_name)
                if value is None:
                    continue
                values = value if isinstance(value, list) else [value]
                references.extend(str(entry) for entry in values if entry)
        return references

    @staticmethod
    def _require_reference_arrays(
        items: list[dict[str, object]], *field_names: str
    ) -> None:
        for index, item in enumerate(items):
            for field_name in field_names:
                if field_name in item and not isinstance(item[field_name], list):
                    raise ArtifactValidationError(
                        "Report IR reference field must be an array: "
                        f"{field_name} at index {index}"
                    )

    @staticmethod
    def _assert_references(
        label: str, references: list[str], known_ids: set[str]
    ) -> None:
        missing = sorted(set(references) - known_ids)
        if missing:
            raise ArtifactValidationError(
                f"Report IR references unknown {label} IDs: {', '.join(missing)}"
            )

    def validate_json(self, path: str | Path) -> str:
        path = Path(path)
        text = path.read_text(encoding="utf-8")
        self.validate_no_sensitive_text(text, location=str(path))
        json.loads(text)
        return self.sha256(path)

    def validate_text_artifact(self, path: str | Path) -> str:
        path = Path(path)
        text = path.read_text(encoding="utf-8")
        self.validate_no_sensitive_text(text, location=str(path))
        return self.sha256(path)

    def validate_xlsx(self, path: str | Path, *, ir=None) -> str:
        path = Path(path)
        if path.suffix.lower() != ".xlsx":
            raise ArtifactValidationError("Excel artifact must use .xlsx")
        if not path.is_file() or path.stat().st_size == 0:
            raise ArtifactValidationError("XLSX artifact is empty")
        with zipfile.ZipFile(path) as archive:
            macro_entries = [
                name for name in archive.namelist() if name.lower().endswith("vbaproject.bin")
            ]
        if macro_entries:
            raise ArtifactValidationError("XLSX artifact contains VBA project")
        required_sheets = {
            "Executive Summary",
            "Competitor Matrix",
            "Pricing",
            "Market Positioning",
            "Business Model",
            "Trends",
            "Change Events",
            "Data Quality",
            "Evidence",
            "Claims",
            "Sources",
            "Analysis Audit",
        }
        workbook = load_workbook(path, read_only=False, keep_vba=False, data_only=False)
        try:
            missing = sorted(required_sheets - set(workbook.sheetnames))
            if missing:
                raise ArtifactValidationError(
                    f"XLSX missing required sheets: {', '.join(missing)}"
                )
            for sheet in workbook.worksheets:
                if sheet.sheet_state != "visible":
                    raise ArtifactValidationError(
                        f"XLSX contains hidden key sheet: {sheet.title}"
                    )
                if sheet.max_row > 1_048_576:
                    raise ArtifactValidationError(
                        f"XLSX sheet exceeds row limit: {sheet.title}"
                    )
                if sheet.freeze_panes != "A2":
                    raise ArtifactValidationError(
                        f"XLSX sheet is not frozen at header: {sheet.title}"
                    )
                if not sheet.auto_filter.ref:
                    raise ArtifactValidationError(
                        f"XLSX sheet has no auto filter: {sheet.title}"
                    )
            self._require_column(workbook["Evidence"], "evidence_id")
            self._require_column(workbook["Evidence"], "source_id")
            self._require_column(workbook["Sources"], "source_id")
            self._require_column(workbook["Analysis Audit"], "analysis_result_id")
            self._validate_workbook_has_no_sensitive_text(workbook)
            self._validate_numeric_formats(workbook)
            if ir is not None:
                self._validate_analysis_numbers(workbook, ir.analysis_results)
        finally:
            workbook.close()
        return self.sha256(path)

    def _validate_workbook_has_no_sensitive_text(self, workbook) -> None:
        properties = workbook.properties
        property_text = "\n".join(
            str(value)
            for value in (
                properties.title,
                properties.subject,
                properties.creator,
                properties.keywords,
                properties.description,
                properties.lastModifiedBy,
                properties.category,
                properties.contentStatus,
                properties.identifier,
                properties.language,
                properties.version,
                properties.revision,
            )
            if value
        )
        if property_text:
            self.validate_no_sensitive_text(
                property_text, location="XLSX workbook properties"
            )
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value:
                        self.validate_no_sensitive_text(
                            cell.value,
                            location=f"XLSX cell {sheet.title}!{cell.coordinate}",
                        )

    @staticmethod
    def _require_column(sheet, name: str) -> None:
        headers = [cell.value for cell in sheet[1]]
        if name not in headers:
            raise ArtifactValidationError(
                f"XLSX sheet {sheet.title} missing column: {name}"
            )

    @staticmethod
    def _validate_numeric_formats(workbook) -> None:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if (
                        isinstance(cell.value, (int, float))
                        and not isinstance(cell.value, bool)
                        and cell.number_format == "General"
                    ):
                        raise ArtifactValidationError(
                            f"XLSX numeric cell has General format: {sheet.title}!{cell.coordinate}"
                        )

    def _validate_analysis_numbers(self, workbook, analysis_results) -> None:
        sheet = workbook["Analysis Audit"]
        headers = [cell.value for cell in sheet[1]]
        positions = {name: index for index, name in enumerate(headers)}
        for required in ("analysis_result_id", "value_path", "value"):
            if required not in positions:
                raise ArtifactValidationError(
                    f"Analysis Audit missing column: {required}"
                )
        actual: dict[tuple[str, str], list[float]] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            value = row[positions["value"]]
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                continue
            key = (
                str(row[positions["analysis_result_id"]]),
                str(row[positions["value_path"]]),
            )
            actual.setdefault(key, []).append(float(value))
        for result in analysis_results:
            result_id = str(
                result.get("analysis_result_id") or result.get("result_id")
            )
            for root in ("metrics", "tables", "charts"):
                for value_path, value in self._numeric_values(
                    result.get(root), root
                ):
                    candidates = actual.get((result_id, value_path), [])
                    if not any(
                        math.isclose(item, float(value), rel_tol=1e-12, abs_tol=1e-12)
                        for item in candidates
                    ):
                        raise ArtifactValidationError(
                            "XLSX analysis number mismatch: "
                            f"{result_id} {value_path}={value}"
                        )

    @classmethod
    def _numeric_values(cls, value, path: str):
        if isinstance(value, dict):
            rows = []
            for key in sorted(value):
                rows.extend(cls._numeric_values(value[key], f"{path}.{key}"))
            return rows
        if isinstance(value, list):
            rows = []
            for index, item in enumerate(value):
                rows.extend(cls._numeric_values(item, f"{path}[{index}]"))
            return rows
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return [(path, value)]
        return []

    def validate_pdf(
        self,
        path: str | Path,
        *,
        log_path: str | Path | None = None,
        template_version: str | None = None,
        max_output_bytes: int = 100_000_000,
    ) -> str:
        path = Path(path)
        if not path.is_file() or path.stat().st_size == 0:
            raise ArtifactValidationError("PDF artifact is empty")
        if path.stat().st_size > max_output_bytes:
            raise ArtifactValidationError("PDF artifact exceeds maximum size")
        with path.open("rb") as file:
            if file.read(5) != b"%PDF-":
                raise ArtifactValidationError("PDF artifact has invalid header")
        try:
            import fitz

            with fitz.open(path) as document:
                if document.page_count < 1:
                    raise ArtifactValidationError("PDF artifact has no pages")
                metadata_text = "\n".join(
                    str(value)
                    for value in (document.metadata or {}).values()
                    if value
                )
                if metadata_text:
                    self.validate_no_sensitive_text(
                        metadata_text, location=f"{path} metadata"
                    )
                extracted_text_parts = []
                for page in document:
                    try:
                        extracted_text_parts.append(page.get_text("text"))
                    except Exception:
                        continue
                extracted_text = "\n".join(
                    part for part in extracted_text_parts if part
                )
                if extracted_text:
                    self.validate_no_sensitive_text(
                        extracted_text, location=f"{path} text"
                    )
        except ArtifactValidationError:
            raise
        except Exception as exc:
            raise ArtifactValidationError(f"PDF cannot be opened: {exc}") from exc
        if template_version is not None and not template_version.strip():
            raise ArtifactValidationError("PDF template version is missing")
        if log_path is not None:
            log = Path(log_path).read_text(encoding="utf-8", errors="replace")
            self.validate_no_sensitive_text(log, location=str(log_path))
            lowered = log.lower()
            fatal_markers = (
                "fatal error occurred",
                "emergency stop",
                "fontspec error",
                "font not found",
            )
            if any(marker in lowered for marker in fatal_markers):
                raise ArtifactValidationError("PDF log contains fatal error")
        return self.sha256(path)

    def validate_chart_bundle(
        self,
        *,
        png_path: str | Path,
        svg_path: str | Path,
        csv_path: str | Path,
        metadata_path: str | Path,
        dataset_ids: set[str] | None = None,
        analysis_result_ids: set[str] | None = None,
        source_ids: set[str] | None = None,
    ) -> dict[str, str]:
        png_path = Path(png_path)
        svg_path = Path(svg_path)
        csv_path = Path(csv_path)
        metadata_path = Path(metadata_path)
        self._validate_png(png_path)
        self._validate_svg(svg_path)
        rows = self._validate_chart_csv(csv_path)
        metadata = self._validate_chart_metadata(
            metadata_path,
            dataset_ids=dataset_ids,
            analysis_result_ids=analysis_result_ids,
            source_ids=source_ids,
        )
        if not rows:
            raise ArtifactValidationError("Chart data CSV has no rows")
        if metadata["chart_id"] not in png_path.name:
            raise ArtifactValidationError("Chart PNG filename does not contain chart_id")
        return {
            "png": self.sha256(png_path),
            "svg": self.sha256(svg_path),
            "csv": self.sha256(csv_path),
            "metadata": self.sha256(metadata_path),
        }

    def _validate_png(self, path: Path) -> None:
        if not path.is_file() or path.stat().st_size == 0:
            raise ArtifactValidationError("Chart PNG is empty")
        with path.open("rb") as file:
            if file.read(8) != b"\x89PNG\r\n\x1a\n":
                raise ArtifactValidationError("Chart PNG has invalid header")

    def _validate_svg(self, path: Path) -> None:
        if not path.is_file() or path.stat().st_size == 0:
            raise ArtifactValidationError("Chart SVG is empty")
        text = path.read_text(encoding="utf-8", errors="replace")
        self.validate_no_sensitive_text(text, location=str(path))
        if "<svg" not in text[:500].lower():
            raise ArtifactValidationError("Chart SVG has invalid root")
        lowered = text.lower()
        for forbidden in ("<script", "javascript:", "onload=", "onerror="):
            if forbidden in lowered:
                raise ArtifactValidationError(f"Chart SVG contains unsafe content: {forbidden}")

    def _validate_chart_csv(self, path: Path) -> list[dict[str, str]]:
        if not path.is_file() or path.stat().st_size == 0:
            raise ArtifactValidationError("Chart data CSV is empty")
        text = path.read_text(encoding="utf-8")
        self.validate_no_sensitive_text(text, location=str(path))
        reader = csv.DictReader(text.splitlines())
        missing = {"x", "y"} - set(reader.fieldnames or [])
        if missing:
            raise ArtifactValidationError(
                f"Chart data CSV missing columns: {', '.join(sorted(missing))}"
            )
        rows = list(reader)
        for index, row in enumerate(rows, start=2):
            if row.get("x") in {None, ""} or row.get("y") in {None, ""}:
                raise ArtifactValidationError(f"Chart data CSV has empty x/y at row {index}")
            try:
                float(str(row["y"]))
            except ValueError as exc:
                raise ArtifactValidationError(
                    f"Chart data CSV has non-numeric y at row {index}"
                ) from exc
        return rows

    def _validate_chart_metadata(
        self,
        path: Path,
        *,
        dataset_ids: set[str] | None,
        analysis_result_ids: set[str] | None,
        source_ids: set[str] | None,
    ) -> dict[str, object]:
        text = path.read_text(encoding="utf-8")
        self.validate_no_sensitive_text(text, location=str(path))
        metadata = json.loads(text)
        required = {
            "chart_id",
            "chart_type",
            "title",
            "unit",
            "dataset_ids",
            "analysis_result_ids",
            "source_ids",
            "generated_at",
            "generator_version",
        }
        missing = sorted(required - set(metadata))
        if missing:
            raise ArtifactValidationError(
                f"Chart metadata missing fields: {', '.join(missing)}"
            )
        if metadata["chart_type"] == "3d":
            raise ArtifactValidationError("3D charts are not allowed")
        if not str(metadata.get("unit") or "").strip():
            raise ArtifactValidationError("Chart metadata unit is empty")
        linked = (
            list(metadata.get("dataset_ids") or [])
            + list(metadata.get("analysis_result_ids") or [])
            + list(metadata.get("source_ids") or [])
        )
        if not linked:
            raise ArtifactValidationError("Chart metadata has no data lineage references")
        self._validate_reference_subset(
            "dataset_ids", metadata.get("dataset_ids") or [], dataset_ids
        )
        self._validate_reference_subset(
            "analysis_result_ids",
            metadata.get("analysis_result_ids") or [],
            analysis_result_ids,
        )
        self._validate_reference_subset(
            "source_ids", metadata.get("source_ids") or [], source_ids
        )
        return metadata

    @staticmethod
    def _validate_reference_subset(
        field_name: str, values: list[object], allowed: set[str] | None
    ) -> None:
        if allowed is None:
            return
        unknown = sorted(str(value) for value in values if str(value) not in allowed)
        if unknown:
            raise ArtifactValidationError(
                f"Chart metadata has unknown {field_name}: {', '.join(unknown)}"
            )

    def validate_no_sensitive_text(self, text: str, *, location: str = "artifact") -> None:
        patterns = {
            "Authorization header": r"authorization\s*[:=]\s*bearer\s+[A-Za-z0-9._~+/=-]+",
            "API key": r"\b(?:sk|ak|api[_-]?key)[-_][A-Za-z0-9]{12,}\b",
            "App secret": r"\bapp[_-]?secret\s*[:=]\s*['\"]?[A-Za-z0-9._~+/=-]{8,}",
            "temporary credential": r"\b(?:session[_-]?token|temporary[_-]?credential)\s*[:=]",
            "sensitive local path": r"/(?:home|root|etc|var)/(?:[^/\s]+/)*(?:\.ssh|\.aws|\.config|private|secret|token)",
        }
        for label, pattern in patterns.items():
            if re.search(pattern, text, flags=re.IGNORECASE):
                raise ArtifactValidationError(
                    f"{location} contains sensitive content: {label}"
                )

    @staticmethod
    def sha256(path: str | Path) -> str:
        return hashlib.sha256(Path(path).read_bytes()).hexdigest()
