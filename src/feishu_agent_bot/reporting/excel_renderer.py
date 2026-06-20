from __future__ import annotations

import json
import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .artifact_validator import ArtifactValidator
from .models import BuiltArtifact
from .report_ir import ReportIR


REQUIRED_SHEETS = (
    "Executive Summary",
    "Competitor Matrix",
    "Pricing",
    "Market Positioning",
    "Business Model",
    "Trends",
    "Change Events",
    "Data Quality",
    "Evidence",
    "Claims",
    "Sources",
)

_CATEGORY_IDENTITIES = {
    "Competitor Matrix": {"competitor_benchmark", "competitor_matrix_builder", "competitor_matrix"},
    "Pricing": {"pricing_and_packaging", "pricing_normalizer"},
    "Market Positioning": {"market_positioning", "market_share_calculator"},
    "Business Model": {"business_model", "unit_economics_calculator"},
    "Trends": {"trend_and_change", "growth_metrics_calculator", "trend_detector"},
}

_META_COLUMNS = [
    "analysis_result_id",
    "dataset_id",
    "evidence_ids",
    "source_ids",
    "source_locator",
    "data_quality_flag",
]


class ExcelRenderer:
    def __init__(self, validator: ArtifactValidator | None = None):
        self.validator = validator or ArtifactValidator()

    def render(self, ir: ReportIR, path: str | Path) -> BuiltArtifact:
        path = Path(path)
        if path.suffix.lower() != ".xlsx":
            raise ValueError("professional workbook must use .xlsx")
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)

        self._write_sheet(workbook, "Executive Summary", self._summary_rows(ir))
        for sheet_name, identities in _CATEGORY_IDENTITIES.items():
            self._write_sheet(
                workbook,
                sheet_name,
                self._analysis_rows(ir, identities),
                required_columns=_META_COLUMNS,
            )
        self._write_sheet(workbook, "Change Events", ir.change_events)
        self._write_sheet(
            workbook,
            "Data Quality",
            self._quality_rows(ir),
            required_columns=["dataset_id", "data_quality_flag", "quality_warnings"],
        )
        self._write_sheet(
            workbook,
            "Evidence",
            self._evidence_rows(ir),
            required_columns=["evidence_id", "source_id", "source_locator"],
        )
        self._write_sheet(
            workbook,
            "Claims",
            ir.claims,
            required_columns=["claim_id", "supporting_evidence_ids", "contradicting_evidence_ids"],
        )
        self._write_sheet(
            workbook,
            "Sources",
            self._source_rows(ir),
            required_columns=["source_id", "source_locator"],
        )
        self._write_sheet(
            workbook,
            "Analysis Audit",
            self._analysis_audit_rows(ir),
            required_columns=[
                "analysis_result_id",
                "skill_name",
                "tool_name",
                "value_path",
                "value",
                "value_type",
                "input_dataset_ids",
                "input_evidence_ids",
            ],
        )
        self._write_sheet(
            workbook,
            "Report Sections",
            [
                {
                    "section_id": section.section_id,
                    "section_type": section.section_type,
                    "title": section.title,
                    "body": section.body,
                    "claim_ids": section.claim_ids,
                    "analysis_result_ids": section.analysis_result_ids,
                    "table_ids": section.table_ids,
                    "chart_ids": section.chart_ids,
                    "evidence_ids": section.evidence_ids,
                }
                for section in ir.sections
            ],
            required_columns=["section_id", "title", "body"],
        )

        workbook.properties.title = ir.title
        workbook.properties.subject = f"Report {ir.report_id} version {ir.version}"
        workbook.properties.creator = "EviDelta"
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{path.stem}.", suffix=".xlsx", dir=path.parent
        )
        os.close(fd)
        temporary = Path(temp_name)
        try:
            workbook.save(temporary)
            digest = self.validator.validate_xlsx(temporary, ir=ir)
            os.replace(temporary, path)
        finally:
            temporary.unlink(missing_ok=True)
        return BuiltArtifact("xlsx", path, digest)

    @staticmethod
    def _summary_rows(ir: ReportIR) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = [
            {"item_type": "metadata", "title": "Job ID", "content": ir.job_id},
            {"item_type": "metadata", "title": "Report ID", "content": ir.report_id},
            {"item_type": "metadata", "title": "Version", "content": ir.version},
            {"item_type": "metadata", "title": "Title", "content": ir.title},
            {"item_type": "metadata", "title": "Generated At", "content": ir.generated_at},
        ]
        for index, item in enumerate(ir.executive_summary, start=1):
            row = dict(item)
            row.setdefault("item_type", "conclusion")
            row.setdefault("title", f"Conclusion {index}")
            rows.append(row)
        if not ir.executive_summary:
            for section in ir.sections[:5]:
                rows.append(
                    {
                        "item_type": "section_summary",
                        "title": section.title,
                        "content": section.body,
                        "claim_ids": section.claim_ids,
                        "evidence_ids": section.evidence_ids,
                    }
                )
        return rows

    @staticmethod
    def _analysis_rows(
        ir: ReportIR, identities: set[str]
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for result in ir.analysis_results:
            identity = result.get("skill_name") or result.get("tool_name")
            if identity not in identities:
                continue
            result_id = result.get("analysis_result_id") or result.get("result_id")
            base = {
                "analysis_result_id": result_id,
                "evidence_ids": result.get("input_evidence_ids") or [],
                "source_ids": result.get("source_ids") or [],
                "data_quality_flag": ExcelRenderer._quality_flag(result),
            }
            result_tables = result.get("tables") or []
            if not result_tables:
                rows.append({**base, "summary": result.get("summary", ""), **(result.get("metrics") or {})})
                continue
            for table in result_tables:
                if not isinstance(table, dict):
                    rows.append({**base, "value": table})
                    continue
                dataset_id = table.get("dataset_id")
                table_rows = table.get("rows")
                if not isinstance(table_rows, list):
                    table_rows = [table]
                for table_row in table_rows:
                    payload = dict(table_row) if isinstance(table_row, dict) else {"value": table_row}
                    payload = {
                        **base,
                        "dataset_id": dataset_id or payload.get("dataset_id"),
                        "source_locator": table.get("source_locator") or payload.get("source_locator"),
                        **payload,
                    }
                    rows.append(payload)
        if rows:
            return rows
        keywords = {item.lower().replace("_", " ") for item in identities}
        for table in ir.tables:
            label = f"{table.get('title', '')} {table.get('table_id', '')}".lower().replace("_", " ")
            if not any(keyword in label for keyword in keywords):
                continue
            table_rows = table.get("rows") or []
            for row in table_rows:
                if isinstance(row, dict):
                    rows.append(
                        {
                            **{column: None for column in _META_COLUMNS},
                            "dataset_id": table.get("dataset_id"),
                            "source_locator": (table.get("lineage") or {}).get("source_locator"),
                            **row,
                        }
                    )
        return rows

    @staticmethod
    def _quality_rows(ir: ReportIR) -> list[dict[str, object]]:
        rows = []
        for item in ir.data_quality:
            profile = item.get("profile") or {}
            warnings = profile.get("quality_warnings") or []
            rows.append(
                {
                    "dataset_id": item.get("dataset_id"),
                    "name": item.get("name"),
                    "row_count": item.get("row_count"),
                    "column_count": item.get("column_count"),
                    "data_quality_flag": "warning" if warnings else "ok",
                    "quality_warnings": warnings,
                    "numeric_columns": profile.get("numeric_columns") or [],
                    "date_columns": profile.get("likely_time_columns") or [],
                    "duplicate_row_count": profile.get("duplicate_row_count"),
                    "missing_ratios": profile.get("missing_ratios") or {},
                }
            )
        return rows

    @staticmethod
    def _evidence_rows(ir: ReportIR) -> list[dict[str, object]]:
        rows = []
        for item in ir.evidence_references:
            row = dict(item)
            row.setdefault(
                "source_locator",
                f"snapshot:{row['snapshot_id']}" if row.get("snapshot_id") else f"source:{row.get('source_id', '')}",
            )
            rows.append(row)
        return rows

    @staticmethod
    def _source_rows(ir: ReportIR) -> list[dict[str, object]]:
        rows = []
        for item in ir.sources:
            row = dict(item)
            row.setdefault("source_locator", row.get("canonical_url") or row.get("url"))
            rows.append(row)
        return rows

    @staticmethod
    def _analysis_audit_rows(ir: ReportIR) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for result in ir.analysis_results:
            result_id = result.get("analysis_result_id") or result.get("result_id")
            values = []
            for root in ("metrics", "tables", "charts"):
                values.extend(_flatten_scalars(result.get(root), root))
            if not values:
                values = [("summary", result.get("summary", ""))]
            for value_path, value in values:
                rows.append(
                    {
                        "analysis_result_id": result_id,
                        "skill_name": result.get("skill_name"),
                        "tool_name": result.get("tool_name"),
                        "value_path": value_path,
                        "value": value,
                        "value_type": type(value).__name__,
                        "input_dataset_ids": result.get("input_dataset_ids") or [],
                        "input_evidence_ids": result.get("input_evidence_ids") or [],
                        "summary": result.get("summary"),
                        "limitations": result.get("limitations") or [],
                    }
                )
        return rows

    @staticmethod
    def _quality_flag(result: dict) -> str:
        metrics = result.get("metrics") or {}
        if metrics.get("status") == "insufficient_data":
            return "insufficient_data"
        return "warning" if result.get("limitations") else "ok"

    def _write_sheet(
        self,
        workbook: Workbook,
        name: str,
        rows: Iterable[dict[str, object]],
        *,
        required_columns: list[str] | None = None,
    ) -> None:
        normalized = [dict(row) for row in rows if isinstance(row, dict)]
        columns = list(required_columns or [])
        for row in normalized:
            for key in row:
                if key not in columns:
                    columns.append(str(key))
        if not columns:
            columns = ["status"]
            normalized = [{"status": "no_data"}]
        sheet = workbook.create_sheet(name)
        sheet.append(columns)
        for row in normalized:
            sheet.append([_excel_value(row.get(column), column) for column in columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 24
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                self._apply_number_format(cell, columns[cell.column - 1])
        for index, column in enumerate(columns, start=1):
            values = [str(column)] + [str(sheet.cell(row=row, column=index).value or "") for row in range(2, min(sheet.max_row, 200) + 1)]
            width = min(60, max(12, max((len(value) for value in values), default=12) + 2))
            sheet.column_dimensions[get_column_letter(index)].width = width

    @staticmethod
    def _apply_number_format(cell, column_name: str) -> None:
        if not isinstance(cell.value, (int, float)) or isinstance(cell.value, bool):
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = "yyyy-mm-dd"
            return
        name = column_name.lower()
        if any(token in name for token in ("percent", "percentage", "ratio", "share", "growth", "completeness")) and abs(cell.value) <= 1:
            cell.number_format = "0.00%"
        elif any(token in name for token in ("price", "cost", "revenue", "amount", "currency", "价格", "费用", "收入")):
            cell.number_format = '#,##0.00;[Red]-#,##0.00'
        else:
            cell.number_format = "#,##0.00" if isinstance(cell.value, float) else "#,##0"


def _excel_value(value, column_name: str = ""):
    if isinstance(value, str) and any(
        token in column_name.lower()
        for token in ("date", "time", "published_at", "observed_at", "effective_at", "created_at")
    ):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(
                tzinfo=None
            )
        except ValueError:
            pass
    if value is None or isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _flatten_scalars(value, path: str):
    if isinstance(value, dict):
        result = []
        for key in sorted(value):
            result.extend(_flatten_scalars(value[key], f"{path}.{key}"))
        return result
    if isinstance(value, list):
        result = []
        for index, item in enumerate(value):
            result.extend(_flatten_scalars(item, f"{path}[{index}]"))
        return result
    if value is None:
        return []
    return [(path, value)]
