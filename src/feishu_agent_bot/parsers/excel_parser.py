from __future__ import annotations

from datetime import date, datetime
import importlib.util
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .base import ParsedAsset, ParsedTable
from .exceptions import ExcelParseError
from .tabular_utils import normalize_scalar, unique_headers


class ExcelParser:
    name = "excel"
    version = "2.0"
    supported_file_types = {"excel"}
    supported_mime_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/vnd.ms-excel.sheet.binary.macroenabled.12",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }

    def can_parse(self, asset) -> bool:
        return (
            asset.file_type in self.supported_file_types
            or asset.detected_mime_type in self.supported_mime_types
        )

    def parse(self, path: Path, *, asset_id: str) -> ParsedAsset:
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return self._parse_openxml(path, asset_id=asset_id)
        if suffix == ".xls":
            return self._parse_legacy(path, asset_id=asset_id, engine="xlrd")
        if suffix == ".xlsb":
            return self._parse_legacy(path, asset_id=asset_id, engine="pyxlsb")
        raise ExcelParseError(f"不支持的 Excel 扩展名: {suffix or '(none)'}")

    def _parse_openxml(self, path: Path, *, asset_id: str) -> ParsedAsset:
        warnings: list[str] = []
        if path.suffix.lower() in {".xlsm", ".xltm"}:
            warnings.append("macro_content_ignored; VBA was not loaded or executed")
        try:
            values_book = load_workbook(
                path,
                read_only=False,
                data_only=True,
                keep_vba=False,
                keep_links=False,
            )
            formula_book = load_workbook(
                path,
                read_only=False,
                data_only=False,
                keep_vba=False,
                keep_links=False,
            )
        except Exception as exc:
            raise ExcelParseError(f"Excel OpenXML 解析失败: {exc}") from exc

        calculation = self._calculation_metadata(formula_book)
        named_ranges = self._named_ranges(formula_book)
        tables: list[ParsedTable] = []
        formula_count = 0
        try:
            for index, sheet_name in enumerate(formula_book.sheetnames, start=1):
                formula_sheet = formula_book[sheet_name]
                value_sheet = values_book[sheet_name]
                bounds = self._used_bounds(formula_sheet)
                if bounds is None:
                    continue
                min_row, min_col, max_row, max_col = bounds
                matrix = self._expanded_matrix(
                    value_sheet, min_row, min_col, max_row, max_col
                )
                formula_matrix = self._expanded_matrix(
                    formula_sheet, min_row, min_col, max_row, max_col
                )
                header_rows = self._header_row_count(matrix, formula_sheet, min_row)
                raw_headers = self._flatten_headers(matrix[:header_rows])
                headers, header_warnings = unique_headers(raw_headers)
                warnings.extend(f"sheet={sheet_name}; {item}" for item in header_warnings)
                records: list[dict[str, object]] = []
                cell_metadata: list[dict[str, object]] = []
                for row_offset, row in enumerate(matrix[header_rows:], start=header_rows):
                    record: dict[str, object] = {}
                    for column_offset, column in enumerate(headers):
                        value = row[column_offset] if column_offset < len(row) else None
                        record[column] = normalize_scalar(value)
                        source_row = min_row + row_offset
                        source_col = min_col + column_offset
                        coordinate = f"{get_column_letter(source_col)}{source_row}"
                        formula = formula_matrix[row_offset][column_offset]
                        source_cell = formula_sheet.cell(source_row, source_col)
                        if isinstance(formula, str) and formula.startswith("="):
                            formula_count += 1
                            cell_metadata.append(
                                {
                                    "asset_id": asset_id,
                                    "sheet_name": sheet_name,
                                    "cell_range": coordinate,
                                    "column": column,
                                    "formula": formula,
                                    "cached_value": normalize_scalar(value),
                                    "calculation_status": "stale_or_unknown",
                                    "number_format": source_cell.number_format,
                                }
                            )
                        elif self._has_semantic_format(source_cell.number_format, value):
                            cell_metadata.append(
                                {
                                    "asset_id": asset_id,
                                    "sheet_name": sheet_name,
                                    "cell_range": coordinate,
                                    "column": column,
                                    "formula": None,
                                    "cached_value": normalize_scalar(value),
                                    "calculation_status": "not_applicable",
                                    "number_format": source_cell.number_format,
                                }
                            )
                    records.append(record)
                cell_range = (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(max_col)}{max_row}"
                )
                locator = (
                    f"{path.name}#sheet={quote(sheet_name, safe='')}"
                    f"&range={cell_range}"
                )
                tables.append(
                    ParsedTable(
                        f"{asset_id}-T{index:03d}",
                        headers,
                        records,
                        caption=sheet_name,
                        sheet_name=sheet_name,
                        cell_range=cell_range,
                        source_locator=locator,
                        extraction_method="excel_sheet",
                        metadata={
                            "sheet_state": formula_sheet.sheet_state,
                            "header_row_count": header_rows,
                            "merged_ranges": [
                                str(item) for item in formula_sheet.merged_cells.ranges
                            ],
                            "cell_metadata": cell_metadata,
                            "named_ranges": [
                                item
                                for item in named_ranges
                                if item.get("sheet_name") in {None, sheet_name}
                            ],
                            "calculation": calculation,
                        },
                    )
                )
        finally:
            values_book.close()
            formula_book.close()
        if formula_count:
            warnings.append(
                f"formula_cells={formula_count}; cached values are stale_or_unknown"
            )
        metadata = {
            "sheet_count": len(tables),
            "sheets": [
                {
                    "name": table.sheet_name,
                    "state": table.metadata.get("sheet_state"),
                    "cell_range": table.cell_range,
                }
                for table in tables
            ],
            "named_ranges": named_ranges,
            "formula_count": formula_count,
            "calculation": calculation,
            "macros_loaded": False,
            "external_links_loaded": False,
        }
        return ParsedAsset(
            asset_id,
            "excel",
            path.stem,
            tables=tables,
            metadata=metadata,
            warnings=warnings,
            extraction_method="excel_openxml",
        )

    def _parse_legacy(
        self, path: Path, *, asset_id: str, engine: str
    ) -> ParsedAsset:
        if importlib.util.find_spec(engine) is None:
            raise ExcelParseError(
                f"解析 {path.suffix} 需要可选依赖 {engine}，当前环境未安装"
            )
        import pandas as pd

        try:
            sheets = pd.read_excel(path, sheet_name=None, header=None, engine=engine)
        except Exception as exc:
            raise ExcelParseError(f"Excel {path.suffix} 解析失败: {exc}") from exc
        tables: list[ParsedTable] = []
        warnings = [
            f"{path.suffix} parser does not expose formulas, merged cells, or named ranges"
        ]
        for index, (sheet_name, frame) in enumerate(sheets.items(), start=1):
            matrix = frame.where(frame.notna(), None).values.tolist()
            if not matrix:
                continue
            header_rows = self._header_row_count(matrix, None, 1)
            headers, header_warnings = unique_headers(
                self._flatten_headers(matrix[:header_rows])
            )
            warnings.extend(f"sheet={sheet_name}; {item}" for item in header_warnings)
            records = [
                {
                    header: normalize_scalar(row[column_index])
                    if column_index < len(row)
                    else None
                    for column_index, header in enumerate(headers)
                }
                for row in matrix[header_rows:]
            ]
            max_row, max_col = len(matrix), len(headers)
            cell_range = f"A1:{get_column_letter(max_col)}{max_row}"
            tables.append(
                ParsedTable(
                    f"{asset_id}-T{index:03d}",
                    headers,
                    records,
                    caption=str(sheet_name),
                    sheet_name=str(sheet_name),
                    cell_range=cell_range,
                    source_locator=(
                        f"{path.name}#sheet={quote(str(sheet_name), safe='')}"
                        f"&range={cell_range}"
                    ),
                    extraction_method=f"excel_{engine}",
                    metadata={
                        "sheet_state": "unknown",
                        "header_row_count": header_rows,
                        "calculation": {"status": "stale_or_unknown"},
                    },
                )
            )
        return ParsedAsset(
            asset_id,
            "excel",
            path.stem,
            tables=tables,
            metadata={
                "sheet_count": len(tables),
                "engine": engine,
                "macros_loaded": False,
                "external_links_loaded": False,
            },
            warnings=warnings,
            extraction_method=f"excel_{engine}",
        )

    @staticmethod
    def _used_bounds(sheet) -> tuple[int, int, int, int] | None:
        populated = [
            (cell.row, cell.column)
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        if not populated:
            return None
        rows = [item[0] for item in populated]
        columns = [item[1] for item in populated]
        return min(rows), min(columns), max(rows), max(columns)

    @staticmethod
    def _expanded_matrix(sheet, min_row: int, min_col: int, max_row: int, max_col: int) -> list[list[object]]:
        merged_values: dict[tuple[int, int], object] = {}
        for merged in sheet.merged_cells.ranges:
            value = sheet.cell(merged.min_row, merged.min_col).value
            for row in range(merged.min_row, merged.max_row + 1):
                for column in range(merged.min_col, merged.max_col + 1):
                    merged_values[(row, column)] = value
        return [
            [
                merged_values.get((row, column), sheet.cell(row, column).value)
                for column in range(min_col, max_col + 1)
            ]
            for row in range(min_row, max_row + 1)
        ]

    @staticmethod
    def _header_row_count(matrix: list[list[object]], sheet, min_row: int) -> int:
        if len(matrix) <= 1:
            return 1
        merged_header_end = 1
        if sheet is not None:
            for merged in sheet.merged_cells.ranges:
                if merged.min_row <= min_row + 2 and merged.min_row >= min_row:
                    merged_header_end = max(
                        merged_header_end, merged.max_row - min_row + 1
                    )
        # A row dominated by numbers/dates is the first data row. Limit header
        # discovery to three rows to avoid consuming sparse worksheets as headers.
        for index, row in enumerate(matrix[1:3], start=1):
            populated = [value for value in row if value not in {None, ""}]
            numeric = [
                value
                for value in populated
                if isinstance(value, (int, float, date, datetime))
                and not isinstance(value, bool)
            ]
            if populated and len(numeric) / len(populated) >= 0.5:
                return max(1, min(index, merged_header_end))
        return min(max(1, merged_header_end), min(3, len(matrix) - 1))

    @staticmethod
    def _flatten_headers(rows: list[list[object]]) -> list[str]:
        width = max((len(row) for row in rows), default=0)
        result: list[str] = []
        for column in range(width):
            parts: list[str] = []
            for row in rows:
                value = normalize_scalar(row[column]) if column < len(row) else None
                text = "" if value is None else str(value).strip()
                if text and (not parts or parts[-1] != text):
                    parts.append(text)
            result.append(" / ".join(parts))
        return result

    @staticmethod
    def _named_ranges(book) -> list[dict[str, object]]:
        result: list[dict[str, object]] = []
        for name, defined_name in book.defined_names.items():
            destinations: list[dict[str, str]] = []
            try:
                for sheet_name, cell_range in defined_name.destinations:
                    destinations.append(
                        {"sheet_name": sheet_name, "cell_range": cell_range}
                    )
            except (AttributeError, TypeError, ValueError):
                pass
            if destinations:
                for destination in destinations:
                    result.append({"name": name, **destination})
            else:
                result.append(
                    {
                        "name": name,
                        "sheet_name": None,
                        "cell_range": None,
                        "value": defined_name.attr_text,
                    }
                )
        return result

    @staticmethod
    def _calculation_metadata(book) -> dict[str, object]:
        properties = getattr(book, "calculation", None)
        return {
            "status": "stale_or_unknown",
            "calculation_mode": getattr(properties, "calcMode", None),
            "full_calculation_on_load": getattr(properties, "fullCalcOnLoad", None),
            "force_full_calculation": getattr(properties, "forceFullCalc", None),
        }

    @staticmethod
    def _has_semantic_format(number_format: str, value: object) -> bool:
        lowered = (number_format or "").lower()
        return (
            isinstance(value, (date, datetime))
            or "%" in lowered
            or any(symbol in number_format for symbol in ("$", "¥", "€", "£"))
        )
