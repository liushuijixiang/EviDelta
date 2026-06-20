from __future__ import annotations

from dataclasses import asdict, replace
import json
import threading
import time

from openpyxl import load_workbook
import pytest

from feishu_agent_bot.acquisition import FileTypeDetector
from feishu_agent_bot.analysis import AnalysisExecutor
from feishu_agent_bot.datasets import DatasetProfiler, TabularDataset
from feishu_agent_bot.parsers import default_parser_registry
from feishu_agent_bot.parsers.pdf_parser import PdfParser
from feishu_agent_bot.reporting.artifacts import ProfessionalArtifactBuilder
from feishu_agent_bot.reporting.artifact_validator import (
    ArtifactValidationError,
    ArtifactValidator,
)
from feishu_agent_bot.reporting.excel_renderer import ExcelRenderer
from feishu_agent_bot.reporting.chart_renderer import ChartRenderer
from feishu_agent_bot.reporting.ir_builder import ReportIRBuilder
from feishu_agent_bot.reporting.latex_renderer import LaTeXRenderer
from feishu_agent_bot.reporting.models import BuiltArtifact
from feishu_agent_bot.reporting.report_ir import ReportIR, ReportSection


def valid_report_ir() -> ReportIR:
    return ReportIR(
        job_id="J-VALID",
        report_id="R-VALID",
        title="可追溯报告",
        sections=[
            ReportSection(
                "S1",
                "摘要",
                "正文",
                claim_ids=["C1"],
                evidence_ids=["E1"],
                analysis_result_ids=["AR1"],
                table_ids=["T1"],
                chart_ids=["CH1"],
            )
        ],
        tables=[{"table_id": "T1", "analysis_result_id": "AR1", "rows": []}],
        charts=[
            {
                "chart_id": "CH1",
                "analysis_result_ids": ["AR1"],
                "points": [{"x": "A", "y": 1}],
            }
        ],
        analysis_results=[
            {
                "analysis_result_id": "AR1",
                "input_evidence_ids": ["E1"],
                "input_dataset_ids": [],
            }
        ],
        claims=[
            {
                "claim_id": "C1",
                "supporting_evidence_ids": ["E1"],
                "contradicting_evidence_ids": [],
            }
        ],
        evidence_references=[{"evidence_id": "E1"}],
    )


def test_artifact_validator_accepts_valid_report_ir_and_json(tmp_path):
    ir = valid_report_ir()
    path = tmp_path / "report_ir.json"
    path.write_text(json.dumps(asdict(ir), ensure_ascii=False), encoding="utf-8")

    payload = ArtifactValidator().validate_report_ir(ir)
    digest = ArtifactValidator().validate_report_ir_json(path)

    assert payload["job_id"] == ir.job_id
    assert len(digest) == 64


@pytest.mark.parametrize(
    ("field_name", "expected"),
    [
        ("claim_ids", "unknown claim IDs"),
        ("evidence_ids", "unknown evidence IDs"),
        ("analysis_result_ids", "unknown analysis result IDs"),
        ("table_ids", "unknown table IDs"),
        ("chart_ids", "unknown chart IDs"),
    ],
)
def test_artifact_validator_rejects_broken_section_references(
    field_name, expected
):
    ir = valid_report_ir()
    section = replace(ir.sections[0], **{field_name: ["MISSING"]})

    with pytest.raises(ArtifactValidationError, match=expected):
        ArtifactValidator().validate_report_ir(replace(ir, sections=[section]))


def test_artifact_validator_rejects_duplicate_section_ids():
    ir = valid_report_ir()
    duplicate = ReportSection("S1", "附录", "重复 ID")

    with pytest.raises(ArtifactValidationError, match="duplicate section IDs"):
        ArtifactValidator().validate_report_ir(
            replace(ir, appendices=[duplicate])
        )


def test_artifact_validator_rejects_empty_section_body():
    ir = valid_report_ir()
    section = replace(ir.sections[0], body="")

    with pytest.raises(ArtifactValidationError, match="body must not be empty"):
        ArtifactValidator().validate_report_ir(replace(ir, sections=[section]))


@pytest.mark.parametrize(
    ("collection", "field_name"),
    [
        ("claims", "supporting_evidence_ids"),
        ("analysis_results", "input_evidence_ids"),
        ("change_events", "evidence_ids"),
    ],
)
def test_artifact_validator_rejects_unknown_nested_evidence_references(
    collection, field_name
):
    payload = asdict(valid_report_ir())
    if collection == "change_events":
        payload[collection] = [{"event_id": "EV1", field_name: ["MISSING"]}]
    else:
        payload[collection][0][field_name] = ["MISSING"]

    with pytest.raises(ArtifactValidationError, match="unknown evidence IDs"):
        ArtifactValidator().validate_report_ir(payload)


@pytest.mark.parametrize("collection", ["tables", "charts"])
def test_artifact_validator_rejects_unknown_analysis_result_references(collection):
    payload = asdict(valid_report_ir())
    if collection == "tables":
        payload[collection][0]["analysis_result_id"] = "MISSING"
    else:
        payload[collection][0]["analysis_result_ids"] = ["MISSING"]

    with pytest.raises(
        ArtifactValidationError, match="unknown analysis result IDs"
    ):
        ArtifactValidator().validate_report_ir(payload)


def test_artifact_validator_rejects_invalid_report_ir_schema():
    payload = asdict(valid_report_ir())
    payload["sections"] = ["not-an-object"]

    with pytest.raises(ArtifactValidationError, match="must be an object"):
        ArtifactValidator().validate_report_ir(payload)


def test_professional_builder_stops_before_writing_invalid_report_ir(tmp_path):
    ir = valid_report_ir()
    invalid = replace(
        ir,
        sections=[replace(ir.sections[0], evidence_ids=["MISSING"])],
    )

    with pytest.raises(ArtifactValidationError, match="unknown evidence IDs"):
        ProfessionalArtifactBuilder(tmp_path).build(invalid, ["json"])

    assert not (tmp_path / ir.job_id).exists()


def test_professional_builder_renders_pdf_and_xlsx_with_bounded_concurrency(
    tmp_path
):
    lock = threading.Lock()
    current = 0
    peak = 0

    def write_artifact(path, artifact_type):
        nonlocal current, peak
        with lock:
            current += 1
            peak = max(peak, current)
        time.sleep(0.05)
        path.write_bytes(artifact_type.encode("ascii"))
        with lock:
            current -= 1
        return BuiltArtifact(artifact_type, path, artifact_type + "-hash")

    class FakeExcelRenderer:
        def render(self, ir, path):
            return write_artifact(path, "xlsx")

    class FakeLatexRenderer:
        def render(self, ir):
            return "report"

        def render_bibliography(self, ir):
            return ""

    class FakePDFCompiler:
        def compile(self, latex, path, **kwargs):
            return write_artifact(path, "pdf")

    ir = ReportIR(job_id="J-PARALLEL", report_id="R1", title="并发", sections=[])
    builder = ProfessionalArtifactBuilder(tmp_path, max_concurrency=2)
    builder.excel_renderer = FakeExcelRenderer()
    builder.latex_renderer = FakeLatexRenderer()
    builder.pdf_compiler = FakePDFCompiler()

    artifacts = builder.build(ir, ["xlsx", "pdf"])

    assert peak == 2
    types = [artifact.artifact_type for artifact in artifacts]
    assert types.index("xlsx") < types.index("pdf")


def test_pdf_recompiles_when_latex_or_chart_input_changes(tmp_path):
    class FakePDFCompiler:
        max_output_bytes = 100_000_000

        def __init__(self):
            self.calls = 0

        def compile(self, _latex, path, **_kwargs):
            self.calls += 1
            path.write_bytes(f"pdf-{self.calls}".encode("ascii"))
            return BuiltArtifact("pdf", path, f"hash-{self.calls}")

    builder = ProfessionalArtifactBuilder(tmp_path)
    compiler = FakePDFCompiler()
    builder.pdf_compiler = compiler
    builder.validator.validate_pdf = lambda *_args, **_kwargs: "valid-hash"
    first = ReportIR(
        job_id="J-PDF-INPUT",
        report_id="R1",
        title="第一版标题",
        sections=[ReportSection("S1", "摘要", "第一版正文")],
    )
    changed = replace(
        first,
        title="第二版标题",
        sections=[ReportSection("S1", "摘要", "第二版正文")],
    )

    builder.compile_pdf(first)
    builder.compile_pdf(first)
    builder.compile_pdf(changed)

    assert compiler.calls == 2
    assert (
        builder.job_dir(first) / "report.pdf.input.sha256"
    ).read_text(encoding="ascii")


def test_report_ir_builder_enforces_chart_and_table_limits(tmp_path):
    report_json = tmp_path / "report.json"
    report_json.write_text(
        json.dumps({"sections": [{"section_id": "S1", "title": "摘要"}]}),
        encoding="utf-8",
    )
    datasets = [
        {
            "dataset_id": f"D{index}",
            "table_id": f"T{index}",
            "rows": [{"value": index}],
        }
        for index in range(3)
    ]
    analysis_results = [
        {
            "analysis_result_id": "AR1",
            "tool_name": "test",
            "charts": [
                {"chart_id": f"CH{index}", "points": [{"x": "A", "y": index}]}
                for index in range(3)
            ],
        }
    ]

    ir = ReportIRBuilder(max_tables=2, max_charts=1).from_report_json(
        job_id="J1",
        topic="limits",
        report_json_path=report_json,
        datasets=datasets,
        analysis_results=analysis_results,
    )

    assert len(ir.tables) == 2
    assert len(ir.charts) == 1
    assert any("表格数量" in item for item in ir.limitations)
    assert any("图表数量" in item for item in ir.limitations)


def test_file_type_detector_uses_extension_content_type_and_magic():
    assert FileTypeDetector.detect("prices.csv") == "csv"
    assert FileTypeDetector.detect("x", content_type="application/pdf") == "pdf"
    assert FileTypeDetector.detect("x", sample=b"%PDF-1.7") == "pdf"


def test_parser_registry_parses_html_csv_json_and_text(tmp_path):
    registry = default_parser_registry()
    html = tmp_path / "sample_page.html"
    html.write_text(
        "<html><body><h1>竞品</h1><table><tr><th>公司</th><th>价格</th></tr>"
        "<tr><td>A</td><td>99</td></tr></table></body></html>",
        encoding="utf-8",
    )
    csv = tmp_path / "sample_prices.csv"
    csv.write_text("company,price\nA,99\nB,129\n", encoding="utf-8")
    data = tmp_path / "sample_metrics.json"
    data.write_text(
        json.dumps({"rows": [{"metric": "growth", "value": 0.2}]}),
        encoding="utf-8",
    )
    text = tmp_path / "sample.txt"
    text.write_text("plain evidence", encoding="utf-8")

    parsed_html = registry.parse("html", html, asset_id="A1")
    parsed_csv = registry.parse("csv", csv, asset_id="A2")
    parsed_json = registry.parse("json", data, asset_id="A3")
    parsed_text = registry.parse("text", text, asset_id="A4")

    assert parsed_html.tables[0].rows[0]["公司"] == "A"
    assert parsed_csv.tables[0].rows[1]["price"] == 129
    assert parsed_json.tables[0].rows[0]["metric"] == "growth"
    assert parsed_text.text_blocks[0].text == "plain evidence"


def test_pdf_parser_extracts_text_blocks_page_metadata_and_locator(tmp_path):
    import fitz

    pdf_path = tmp_path / "sample_text.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 72), "PDF evidence quote")
    document.set_metadata({"title": "Sample PDF"})
    document.save(pdf_path)
    document.close()

    parsed = PdfParser().parse(pdf_path, asset_id="PDF1")

    assert parsed.file_type == "pdf"
    assert parsed.title == "Sample PDF"
    assert parsed.metadata["page_count"] == 1
    assert parsed.metadata["parser"] == "PyMuPDF"
    assert parsed.text_blocks[0].page_number == 1
    assert parsed.text_blocks[0].bbox is not None
    assert "PDF evidence quote" in parsed.text_blocks[0].text
    assert parsed.text_blocks[0].source_locator.startswith("sample_text.pdf#page=1")


def test_pdf_parser_extracts_table_and_image_metadata(tmp_path):
    import fitz

    pdf_path = tmp_path / "sample_table.pdf"
    document = fitz.open()
    page = document.new_page(width=420, height=320)
    x_positions = [50, 180, 300]
    y_positions = [50, 90, 130]
    for x in x_positions:
        page.draw_line((x, 50), (x, 130), color=(0, 0, 0), width=1)
    for y in y_positions:
        page.draw_line((50, y), (300, y), color=(0, 0, 0), width=1)
    for x, y, text in (
        (60, 75, "Company"),
        (190, 75, "Price"),
        (60, 115, "Alpha"),
        (190, 115, "99"),
    ):
        page.insert_text((x, y), text, fontsize=10)
    pixmap = fitz.Pixmap(fitz.csRGB, (0, 0, 10, 10), False)
    pixmap.clear_with(128)
    page.insert_image(fitz.Rect(330, 50, 380, 100), pixmap=pixmap)
    document.save(pdf_path)
    document.close()

    parsed = PdfParser(ocr_enabled=False).parse(pdf_path, asset_id="PDF-TABLE")

    assert parsed.metadata["table_count"] == 1
    assert parsed.metadata["image_count"] == 1
    table = parsed.tables[0]
    assert table.columns == ["Company", "Price"]
    assert table.rows == [{"Company": "Alpha", "Price": 99}]
    assert table.page_number == 1
    assert table.extraction_method == "pymupdf_find_tables"
    assert table.source_locator == "sample_table.pdf#page=1&table=1"
    assert table.metadata["bbox"] == [50.0, 50.0, 300.0, 130.0]
    image = parsed.metadata["images"][0]
    assert image["page_number"] == 1
    assert image["width"] == 10
    assert image["height"] == 10
    assert image["source_locator"] == "sample_table.pdf#page=1&image=1"
    assert image["placements"] == [[330.0, 50.0, 380.0, 100.0]]


def test_pdf_parser_enforces_parse_timeout(tmp_path, monkeypatch):
    import fitz
    import feishu_agent_bot.parsers.pdf_parser as pdf_parser_module
    from feishu_agent_bot.parsers.exceptions import PdfParseError

    pdf_path = tmp_path / "timeout.pdf"
    document = fitz.open()
    document.new_page()
    document.save(pdf_path)
    document.close()
    clock = iter([0.0, 2.0])
    monkeypatch.setattr(pdf_parser_module, "monotonic", lambda: next(clock))

    with pytest.raises(PdfParseError, match="timed out"):
        PdfParser(timeout_seconds=1).parse(pdf_path, asset_id="PDF-TIMEOUT")


def test_pdf_parser_records_ocr_warning_when_text_is_insufficient(tmp_path):
    import fitz

    pdf_path = tmp_path / "sample_scanned.pdf"
    document = fitz.open()
    page = document.new_page()
    pixmap = fitz.Pixmap(fitz.csRGB, (0, 0, 10, 10), False)
    pixmap.clear_with(255)
    page.insert_image(fitz.Rect(72, 72, 160, 160), pixmap=pixmap)
    document.save(pdf_path)
    document.close()

    parsed = PdfParser(min_text_chars_per_page=50).parse(pdf_path, asset_id="PDF2")

    if parsed.extraction_method == "ocr":
        assert parsed.metadata["original_image_count"] == 1
        assert parsed.metadata["ocr_languages"] == "chi_sim+eng"
        assert any("OCR text defaults" in warning for warning in parsed.warnings)
    else:
        assert parsed.metadata["image_count"] == 1
        assert parsed.text_blocks == []
        assert any("OCR fallback required" in warning for warning in parsed.warnings)


def test_dataset_profiler_and_analysis_executor_are_deterministic():
    dataset = TabularDataset(
        dataset_id="D1",
        job_id="J1",
        asset_id="A1",
        table_id="T1",
        name="competitor pricing",
        columns=["company", "price"],
        rows=[
            {"company": "A", "price": "99"},
            {"company": "B", "price": "129"},
            {"company": "B", "price": "129"},
        ],
    )
    profile = DatasetProfiler().profile(dataset)
    run, results = AnalysisExecutor().run(
        job_id="J1",
        topic="竞品价格趋势",
        datasets=[dataset],
        profiles=[profile],
    )

    assert profile.duplicate_row_count == 1
    assert "price" in profile.numeric_columns
    assert "competitor_matrix" in run.selected_tools
    assert "pricing_normalizer" in run.selected_tools
    assert any(result.tool_name == "data_quality_summarizer" for result in results)


def test_dataset_profiler_builds_quality_report():
    dataset = TabularDataset(
        dataset_id="D2",
        job_id="J1",
        asset_id="A1",
        table_id="T2",
        name="metrics",
        columns=["id", "date", "value", "sparse"],
        rows=[
            {"id": "1", "date": "2026-01", "value": "10", "sparse": ""},
            {"id": "2", "date": "2026-02", "value": "20", "sparse": None},
            {"id": "3", "date": "2026-03", "value": "30", "sparse": ""},
        ],
    )

    profiler = DatasetProfiler()
    profile = profiler.profile(dataset)
    report = profiler.quality_report(profile)

    assert profile.unique_counts["id"] == 3
    assert profile.mean_values["value"] == 20
    assert profile.likely_primary_keys == ["id"]
    assert profile.likely_time_columns == ["date"]
    assert report.quality_level == "low"
    assert "trend" in report.not_usable_for


def test_dataset_profiler_infers_schema_units_currency_quantiles_and_outliers():
    dataset = TabularDataset(
        dataset_id="D3",
        job_id="J1",
        asset_id="A1",
        table_id="T3",
        name="pricing",
        columns=["plan_id", "month", "price_cny", "growth_pct", "latency_ms"],
        rows=[
            {"plan_id": "p1", "month": "2026-01", "price_cny": "¥100", "growth_pct": "5%", "latency_ms": "10ms"},
            {"plan_id": "p2", "month": "2026-02", "price_cny": "¥110", "growth_pct": "6%", "latency_ms": "11ms"},
            {"plan_id": "p3", "month": "2026-03", "price_cny": "¥120", "growth_pct": "7%", "latency_ms": "12ms"},
            {"plan_id": "p4", "month": "2026-04", "price_cny": "¥5000", "growth_pct": "-2%", "latency_ms": "200ms"},
        ],
        lineage={"table_metadata": {"encoding": "utf-8", "delimiter": ","}},
    )

    profile = DatasetProfiler().profile(dataset)
    report = DatasetProfiler().quality_report(profile)
    schema = {column.name: column for column in profile.schema}

    assert schema["price_cny"].inferred_type == "integer"
    assert schema["price_cny"].currency == "CNY"
    assert schema["growth_pct"].unit == "percent"
    assert schema["latency_ms"].unit == "ms"
    assert profile.time_ranges["month"] == {"min": "2026-01-01", "max": "2026-04-01"}
    assert profile.quantiles["price_cny"]["p50"] == 115
    assert profile.negative_ratios["growth_pct"] == 0.25
    assert profile.outlier_candidates["price_cny"][0]["value"] == 5000
    assert "encoding=utf-8" in profile.parsing_warnings
    assert any(issue.issue_type == "outlier_candidates" for issue in report.issues)


def test_report_ir_builder_includes_analysis_results_and_datasets(tmp_path):
    report_json = tmp_path / "report.json"
    report_json.write_text(
        json.dumps(
            {
                "sections": [
                    {"section_id": "S1", "title": "摘要", "body": "报告正文"}
                ]
            }
        ),
        encoding="utf-8",
    )

    ir = ReportIRBuilder().from_report_json(
        job_id="J1",
        topic="竞品价格",
        report_json_path=report_json,
        analysis_results=[
            {
                "analysis_result_id": "AR1",
                "tool_name": "data_quality_summarizer",
                "summary": "数据集数量 1。",
                "metrics": {"dataset_count": 1},
            }
        ],
        datasets=[
            {
                "dataset_id": "D1",
                "table_id": "T1",
                "dataset_name": "prices",
                "columns": ["company", "price"],
                "rows": [{"company": "A", "price": 99}],
                "row_count": 1,
                "column_count": 2,
                "profile": {"numeric_columns": ["price"]},
            }
        ],
        analysis_runs=[
            {
                "analysis_run_id": "RUN1",
                "analysis_plan": {
                    "selected_skills": ["pricing_and_packaging"],
                    "skipped_skills": [
                        {
                            "skill_name": "business_model",
                            "reason": "缺少适用数据集或证据",
                            "missing_inputs": ["business_model_inputs"],
                        }
                    ],
                    "expected_outputs": ["数据质量摘要", "定价与套餐分析"],
                    "limitations": ["未发现商业模式输入"],
                },
            }
        ],
    )

    assert any(section.section_type == "analysis" for section in ir.sections)
    analysis_section = next(
        section for section in ir.sections if section.section_type == "analysis"
    )
    assert analysis_section.chart_ids == [ir.charts[0]["chart_id"]]
    assert "T1" in analysis_section.table_ids
    assert ir.tables[0]["dataset_id"] == "D1"
    assert ir.data_quality[0]["profile"]["numeric_columns"] == ["price"]
    assert ir.charts[0]["analysis_result_ids"] == ["AR1"]
    assert ir.metadata["analysis_result_count"] == 1
    assert ir.metadata["analysis_plan"]["selected_skills"] == ["pricing_and_packaging"]
    assert "未发现商业模式输入" in ir.limitations
    assert any("business_model: 缺少适用数据集或证据" in item for item in ir.limitations)


def test_report_ir_builder_generates_and_binds_evidence_overview_chart(tmp_path):
    report_json = tmp_path / "report.json"
    report_json.write_text(
        json.dumps(
            {
                "language": "zh",
                "sections": [
                    {"section_id": "summary", "title": "摘要", "body": "正文"}
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    ir = ReportIRBuilder().from_report_json(
        job_id="J1",
        topic="测试",
        report_json_path=report_json,
        evidence=[
            {
                "evidence_id": "E1",
                "source_id": "S1",
                "evidence_type": "price",
            },
            {
                "evidence_id": "E2",
                "source_id": "S1",
                "evidence_type": "price",
            },
            {
                "evidence_id": "E3",
                "source_id": "S2",
                "evidence_type": "product_feature",
            },
        ],
    )

    assert ir.charts[0]["chart_id"] == "research-evidence-distribution"
    assert ir.charts[0]["chart_type"] == "pie"
    overview = next(
        section for section in ir.sections
        if section.section_id == "research_data_visualization"
    )
    assert "research-evidence-distribution" in overview.chart_ids
    latex = LaTeXRenderer().render(ir)
    assert "charts/research-evidence-distribution.png" in latex
    assert r"\ref{fig:research-evidence-distribution}" in latex
    assert r"\label{fig:research-evidence-distribution}" in latex


def test_report_ir_builder_derives_comparable_evidence_charts_and_tables(
    tmp_path
):
    report_json = tmp_path / "report.json"
    report_json.write_text(
        json.dumps(
            {
                "language": "zh",
                "sections": [
                    {"section_id": "summary", "title": "摘要", "body": "正文"}
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    evidence = [
        {
            "evidence_id": "E1",
            "source_id": "S1",
            "entity": "企业A",
            "attribute": "市场份额",
            "value": "60%",
            "evidence_type": "metric",
        },
        {
            "evidence_id": "E2",
            "source_id": "S2",
            "entity": "企业B",
            "attribute": "市场份额",
            "value": "40%",
            "evidence_type": "metric",
        },
        {
            "evidence_id": "E3",
            "source_id": "S1",
            "entity": "企业A",
            "attribute": "SWE-bench Verified score",
            "value": "77.2%",
            "evidence_type": "metric",
        },
        {
            "evidence_id": "E4",
            "source_id": "S2",
            "entity": "企业B",
            "attribute": "SWE-bench Verified score",
            "value": "49.0%",
            "evidence_type": "metric",
        },
    ]

    ir = ReportIRBuilder().from_report_json(
        job_id="J1",
        topic="测试",
        report_json_path=report_json,
        evidence=evidence,
        sources=[
            {"source_id": "S1", "title": "来源一"},
            {"source_id": "S2", "title": "来源二"},
        ],
    )

    market_share = next(
        item for item in ir.charts if item["title"] == "市场份额"
    )
    benchmark = next(
        item
        for item in ir.charts
        if item["title"] == "SWE-bench Verified score"
    )
    assert market_share["chart_type"] == "pie"
    assert benchmark["chart_type"] == "bar"
    assert market_share["evidence_ids"] == ["E1", "E2"]
    assert any(
        table["title"] == "市场份额"
        and table["rows"][0]["evidence_id"] == "E1"
        for table in ir.tables
    )


def test_report_ir_builder_backfills_section_body_from_claim_ids(tmp_path):
    report_json = tmp_path / "report.json"
    report_json.write_text(
        json.dumps(
            {
                "summary": ["主要结论"],
                "sections": [
                    {
                        "section_id": "product_comparison",
                        "title": "竞品对比",
                        "claim_ids": ["C-001"],
                    },
                    {
                        "section_id": "risk",
                        "title": "风险",
                        "claim_ids": [],
                    },
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    ir = ReportIRBuilder().from_report_json(
        job_id="J1",
        topic="竞品价格",
        report_json_path=report_json,
        claims=[
            {
                "claim_id": "C-001",
                "statement": "A 产品价格更低。",
                "supporting_evidence_ids": ["E-001"],
                "contradicting_evidence_ids": [],
            }
        ],
        evidence=[
            {
                "evidence_id": "E-001",
                "source_id": "S-001",
                "exact_quote": "A 产品价格更低。",
            }
        ],
    )

    comparison = ir.sections[0]
    risk = ir.sections[1]
    assert "A 产品价格更低" in comparison.body
    assert comparison.claim_ids == ["C-001"]
    assert comparison.evidence_ids == ["E-001"]
    assert risk.body == "- 暂无充分证据。"
    assert ir.executive_summary == [{"text": "主要结论"}]


def test_professional_artifact_builder_writes_json_xlsx_and_pdf_status(tmp_path):
    ir = ReportIR(
        job_id="J1",
        report_id="R1",
        version=2,
        title="竞品调研",
        sections=[ReportSection("S1", "摘要", "结论正文")],
        analysis_results=[
            {
                "analysis_result_id": "AR1",
                "skill_name": "competitor_benchmark",
                "tool_name": "competitor_matrix_builder",
                "summary": "竞品矩阵",
                "input_dataset_ids": ["D1"],
                "input_evidence_ids": ["E1"],
                "metrics": {"completeness": 0.75},
                "tables": [
                    {
                        "dataset_id": "D1",
                        "name": "competitor_matrix",
                        "rows": [{"company": "A", "feature": "fast"}],
                    }
                ],
                "charts": [],
                "limitations": ["样本有限"],
            },
            {
                "analysis_result_id": "AR2",
                "skill_name": "pricing_and_packaging",
                "tool_name": "pricing_normalizer",
                "summary": "价格分析",
                "input_dataset_ids": ["D2"],
                "input_evidence_ids": ["E2"],
                "metrics": {"parsed_price_count": 2},
                "tables": [
                    {
                        "dataset_id": "D2",
                        "name": "normalized_pricing",
                        "rows": [
                            {"company": "A", "normalized_price": 99.0, "currency": "CNY"},
                            {"company": "B", "normalized_price": 129.0, "currency": "CNY"},
                        ],
                    }
                ],
                "charts": [],
                "limitations": [],
            },
        ],
        change_events=[
            {
                "event_id": "EV1",
                "event_type": "price_change",
                "summary": "价格更新",
                "effective_at": "2026-06-19T00:00:00+00:00",
                "evidence_ids": ["E2"],
            }
        ],
        data_quality=[
            {
                "dataset_id": "D1",
                "name": "competitors",
                "row_count": 1,
                "column_count": 2,
                "profile": {"quality_warnings": ["样本较少"]},
            }
        ],
        evidence_references=[
            {
                "evidence_id": "E1",
                "source_id": "SRC1",
                "snapshot_id": "SN1",
                "exact_quote": "产品支持快充。",
            },
            {
                "evidence_id": "E2",
                "source_id": "SRC1",
                "snapshot_id": "SN1",
                "exact_quote": "套餐价格为 99 元。",
            },
        ],
        claims=[
            {
                "claim_id": "C1",
                "statement": "A 支持快充",
                "supporting_evidence_ids": ["E1"],
                "contradicting_evidence_ids": [],
            }
        ],
        sources=[
            {
                "source_id": "SRC1",
                "title": "官方资料",
                "canonical_url": "https://example.com/report",
            }
        ],
        charts=[
            {
                "chart_id": "price_chart",
                "chart_type": "bar",
                "title": "价格对比",
                "unit": "CNY",
                "points": [{"x": "A", "y": 99}, {"x": "B", "y": 129}],
                "dataset_ids": ["D1"],
                "analysis_result_ids": ["AR1"],
                "source_ids": ["SRC1"],
            }
        ],
    )

    artifacts = ProfessionalArtifactBuilder(tmp_path).build(
        ir, ["json", "xlsx", "pdf"]
    )
    by_type = {artifact.artifact_type: artifact for artifact in artifacts}

    assert by_type["json"].path.is_file()
    report_ir = json.loads(by_type["json"].path.read_text(encoding="utf-8"))
    assert report_ir["report_id"] == "R1"
    assert report_ir["version"] == 2
    assert by_type["xlsx"].path.is_file()
    workbook = load_workbook(by_type["xlsx"].path, read_only=False)
    required = {
        "Executive Summary",
        "Competitor Matrix",
        "Pricing",
        "Market Positioning",
        "Business Model",
        "Trends",
        "Change Events",
        "Data Quality",
        "Evidence",
        "Claims",
        "Sources",
    }
    assert required <= set(workbook.sheetnames)
    assert all(sheet.freeze_panes == "A2" for sheet in workbook.worksheets)
    assert all(sheet.auto_filter.ref for sheet in workbook.worksheets)
    assert all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)
    assert [cell.value for cell in workbook["Evidence"][1]][:3] == [
        "evidence_id",
        "source_id",
        "source_locator",
    ]
    pricing_headers = [cell.value for cell in workbook["Pricing"][1]]
    price_column = pricing_headers.index("normalized_price") + 1
    assert workbook["Pricing"].cell(2, price_column).value == 99
    assert workbook["Pricing"].cell(2, price_column).number_format != "General"
    audit_rows = list(workbook["Analysis Audit"].iter_rows(values_only=True))
    assert ("AR1", "competitor_benchmark", "competitor_matrix_builder", "metrics.completeness", 0.75) == audit_rows[1][:5]
    workbook.close()
    assert by_type["pdf"].status in {"ready", "unavailable"}
    if by_type["pdf"].status == "unavailable":
        assert "xelatex" in by_type["pdf"].error_message
    assert by_type["latex_source"].path.is_file()
    assert by_type["bibliography"].path.name == "reference.bib"
    assert "@misc{" in by_type["bibliography"].path.read_text(encoding="utf-8")
    assert by_type["chart_png"].path.is_file()
    assert by_type["chart_svg"].path.is_file()
    assert by_type["chart_data_csv"].path.is_file()
    assert by_type["chart_metadata_json"].path.is_file()
    metadata = json.loads(by_type["chart_metadata_json"].path.read_text(encoding="utf-8"))
    assert metadata["chart_id"] == "price_chart"
    assert metadata["generator_version"] == "1.2"
    assert len(metadata["input_hash"]) == 64
    assert by_type["manifest"].path.is_file()
    manifest = json.loads(by_type["manifest"].path.read_text(encoding="utf-8"))
    assert manifest["report_version"] == 2
    assert manifest["template_version"] == "business_report:2.0"
    assert all(item["artifact_id"] for item in manifest["artifacts"])
    assert all("byte_size" in item for item in manifest["artifacts"])

    workbook = load_workbook(by_type["xlsx"].path)
    audit = workbook["Analysis Audit"]
    audit.cell(2, 5).value = 0.5
    workbook.save(by_type["xlsx"].path)
    workbook.close()
    with pytest.raises(ArtifactValidationError, match="number mismatch"):
        ArtifactValidator().validate_xlsx(by_type["xlsx"].path, ir=ir)


def test_manifest_does_not_count_stale_file_size_for_failed_artifact(tmp_path):
    ir = valid_report_ir()
    builder = ProfessionalArtifactBuilder(tmp_path)
    job_dir = builder.job_dir(ir)
    stale_pdf = job_dir / "report.pdf"
    stale_pdf.write_bytes(b"%PDF- stale previous output")

    manifest_artifact = builder.write_manifest(
        ir,
        [
            BuiltArtifact(
                "pdf",
                stale_pdf,
                "",
                status="failed",
                error_message="compile failed",
            )
        ],
    )[0]
    manifest = json.loads(manifest_artifact.path.read_text(encoding="utf-8"))
    pdf_entry = next(
        item for item in manifest["artifacts"] if item["artifact_type"] == "pdf"
    )

    assert pdf_entry["status"] == "failed"
    assert pdf_entry["byte_size"] == 0
    assert pdf_entry["sha256"] == ""
    assert pdf_entry["error"] == "compile failed"


def test_chart_renderer_validates_lineage_references(tmp_path):
    renderer = ChartRenderer()
    with pytest.raises(ArtifactValidationError, match="unknown source_ids"):
        renderer.render(
            {
                "chart_id": "bad_source",
                "chart_type": "bar",
                "title": "Bad Source",
                "unit": "CNY",
                "dataset_ids": ["D1"],
                "analysis_result_ids": ["AR1"],
                "source_ids": ["SRC_MISSING"],
            },
            tmp_path,
            chart_id="bad_source",
            chart_type="bar",
            points=[{"x": "A", "y": 1}],
            dataset_ids={"D1"},
            analysis_result_ids={"AR1"},
            source_ids={"SRC1"},
        )


def test_chart_renderer_runs_with_bounded_concurrency_and_stable_order(
    tmp_path, monkeypatch
):
    renderer = ChartRenderer(max_concurrency=2)
    lock = threading.Lock()
    current = 0
    peak = 0
    completed = []

    def fake_render(chart, output_dir, *, chart_id, **kwargs):
        nonlocal current, peak
        with lock:
            current += 1
            peak = max(peak, current)
        time.sleep(0.04 if chart_id != "CH1" else 0.08)
        path = output_dir / f"{chart_id}.png"
        path.write_bytes(chart_id.encode("ascii"))
        with lock:
            current -= 1
        return [BuiltArtifact("chart_png", path, chart_id)]

    monkeypatch.setattr(renderer, "render", fake_render)
    artifacts = renderer.render_all(
        [
            {"chart_id": "CH1", "unit": "count", "points": [{"x": "A", "y": 1}]},
            {"chart_id": "CH2", "unit": "count", "points": [{"x": "A", "y": 2}]},
            {"chart_id": "CH3", "unit": "count", "points": [{"x": "A", "y": 3}]},
        ],
        tmp_path,
        on_chart_completed=completed.append,
    )

    assert peak == 2
    assert [artifact.content_hash for artifact in artifacts] == ["CH1", "CH2", "CH3"]
    assert set(completed) == {"CH1", "CH2", "CH3"}


def test_chart_renderer_reuses_checkpoint_and_only_rebuilds_changed_chart(
    tmp_path, monkeypatch
):
    renderer = ChartRenderer(max_concurrency=1)
    original_plot = renderer._plot
    plotted = []

    def tracked_plot(chart, chart_type, title, unit, rows, png_path, svg_path):
        plotted.append(title)
        original_plot(
            chart, chart_type, title, unit, rows, png_path, svg_path
        )

    monkeypatch.setattr(renderer, "_plot", tracked_plot)
    charts = [
        {
            "chart_id": "CH1",
            "title": "Chart 1",
            "unit": "count",
            "points": [{"x": "A", "y": 1}],
            "source_ids": ["SRC1"],
        },
        {
            "chart_id": "CH2",
            "title": "Chart 2",
            "unit": "count",
            "points": [{"x": "A", "y": 2}],
            "source_ids": ["SRC1"],
        },
    ]
    first_dir = tmp_path / "v1" / "charts"
    second_dir = tmp_path / "v2" / "charts"

    first = renderer.render_all(charts, first_dir)
    second = renderer.render_all(charts, first_dir)
    changed = [dict(charts[0]), {**charts[1], "points": [{"x": "A", "y": 20}]}]
    third = renderer.render_all(changed, second_dir, reuse_dirs=[first_dir])

    assert plotted == ["Chart 1", "Chart 2", "Chart 2"]
    assert [item.content_hash for item in first] == [item.content_hash for item in second]
    assert (second_dir / "CH1.png").read_bytes() == (first_dir / "CH1.png").read_bytes()
    first_hashes = {item.path.name: item.content_hash for item in first}
    third_hashes = {item.path.name: item.content_hash for item in third}
    assert third_hashes["CH1.png"] == first_hashes["CH1.png"]
    assert third_hashes["CH2.png"] != first_hashes["CH2.png"]


def test_artifact_validator_rejects_sensitive_chart_metadata(tmp_path):
    png_path = tmp_path / "secret_chart.png"
    svg_path = tmp_path / "secret_chart.svg"
    csv_path = tmp_path / "secret_chart_data.csv"
    metadata_path = tmp_path / "secret_chart_metadata.json"

    png_path.write_bytes(b"\x89PNG\r\n\x1a\nnot-real-but-header-valid")
    svg_path.write_text("<svg xmlns=\"http://www.w3.org/2000/svg\"></svg>", encoding="utf-8")
    csv_path.write_text("x,y,series\nA,1,\n", encoding="utf-8")
    metadata_path.write_text(
        json.dumps(
            {
                "chart_id": "secret_chart",
                "chart_type": "bar",
                "title": "Secret Chart",
                "unit": "count",
                "dataset_ids": ["D1"],
                "analysis_result_ids": ["AR1"],
                "source_ids": ["SRC1"],
                "generated_at": "2026-06-19T00:00:00+00:00",
                "generator_version": "1.0",
                "debug": "Authorization: Bearer example-token-redacted",
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(ArtifactValidationError, match="sensitive content"):
        ArtifactValidator().validate_chart_bundle(
            png_path=png_path,
            svg_path=svg_path,
            csv_path=csv_path,
            metadata_path=metadata_path,
            dataset_ids={"D1"},
            analysis_result_ids={"AR1"},
            source_ids={"SRC1"},
        )


def test_artifact_validator_rejects_sensitive_pdf_text_and_metadata(tmp_path):
    fitz = pytest.importorskip("fitz")
    text_pdf = tmp_path / "sensitive-text.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 72), "Authorization: Bearer example-token-redacted")
    document.save(text_pdf)
    document.close()

    with pytest.raises(ArtifactValidationError, match="sensitive content"):
        ArtifactValidator().validate_pdf(text_pdf)

    metadata_pdf = tmp_path / "sensitive-metadata.pdf"
    document = fitz.open()
    document.new_page().insert_text((72, 72), "普通报告正文")
    document.set_metadata({"subject": "app_secret=example-redacted-value"})
    document.save(metadata_pdf)
    document.close()

    with pytest.raises(ArtifactValidationError, match="sensitive content"):
        ArtifactValidator().validate_pdf(metadata_pdf)


def test_artifact_validator_rejects_sensitive_text_artifacts(tmp_path):
    latex_path = tmp_path / "report.tex"
    latex_path.write_text(
        r"\section{Authorization: Bearer example-token-redacted}",
        encoding="utf-8",
    )

    with pytest.raises(ArtifactValidationError, match="sensitive content"):
        ArtifactValidator().validate_text_artifact(latex_path)


def test_artifact_validator_rejects_sensitive_xlsx_cells_and_properties(tmp_path):
    ir = valid_report_ir()
    xlsx_path = tmp_path / "report.xlsx"
    ExcelRenderer().render(ir, xlsx_path)

    workbook = load_workbook(xlsx_path)
    workbook["Evidence"].cell(2, 1).value = "Authorization: Bearer example-token-redacted"
    workbook.save(xlsx_path)
    workbook.close()

    with pytest.raises(ArtifactValidationError, match="sensitive content"):
        ArtifactValidator().validate_xlsx(xlsx_path, ir=ir)

    xlsx_path = tmp_path / "report-metadata.xlsx"
    ExcelRenderer().render(ir, xlsx_path)
    workbook = load_workbook(xlsx_path)
    workbook.properties.subject = "app_secret=example-redacted-value"
    workbook.save(xlsx_path)
    workbook.close()

    with pytest.raises(ArtifactValidationError, match="sensitive content"):
        ArtifactValidator().validate_xlsx(xlsx_path, ir=ir)


def test_professional_builder_rejects_sensitive_latex_and_bibliography(tmp_path):
    ir = replace(
        valid_report_ir(),
        sections=[replace(valid_report_ir().sections[0], chart_ids=[])],
        charts=[],
        sources=[
            {
                "source_id": "SRC1",
                "title": "Authorization: Bearer example-token-redacted",
                "canonical_url": "https://example.com/report",
            }
        ],
    )

    with pytest.raises(ArtifactValidationError, match="sensitive content"):
        ProfessionalArtifactBuilder(tmp_path).render_latex(ir)


def test_excel_renderer_maps_each_professional_skill_to_business_sheet(tmp_path):
    mapping = {
        "Competitor Matrix": ("competitor_benchmark", "competitor_matrix_builder"),
        "Pricing": ("pricing_and_packaging", "pricing_normalizer"),
        "Market Positioning": ("market_positioning", "market_share_calculator"),
        "Business Model": ("business_model", "unit_economics_calculator"),
        "Trends": ("trend_and_change", "growth_metrics_calculator"),
    }
    results = []
    for index, (sheet_name, (skill, tool)) in enumerate(mapping.items(), start=1):
        results.append(
            {
                "analysis_result_id": f"AR{index}",
                "skill_name": skill,
                "tool_name": tool,
                "summary": sheet_name,
                "metrics": {"score": index / 10},
                "tables": [
                    {
                        "dataset_id": f"D{index}",
                        "rows": [{"marker": sheet_name, "score": index / 10}],
                    }
                ],
                "charts": [],
                "input_dataset_ids": [f"D{index}"],
                "input_evidence_ids": [f"E{index}"],
                "limitations": [],
            }
        )
    ir = ReportIR(job_id="J1", report_id="R1", title="专业报告", sections=[], analysis_results=results)
    path = tmp_path / "report.xlsx"

    ExcelRenderer().render(ir, path)

    workbook = load_workbook(path)
    for sheet_name in mapping:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        marker_column = headers.index("marker") + 1
        assert sheet.cell(2, marker_column).value == sheet_name
    workbook.close()
