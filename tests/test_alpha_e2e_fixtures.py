from __future__ import annotations

import json
from pathlib import Path
import shutil
import subprocess

import fitz
from openpyxl import load_workbook

from feishu_agent_bot.acquisition import FileTypeDetector
from feishu_agent_bot.datasets import DatasetProfiler, TabularDataset
from feishu_agent_bot.parsers import default_parser_registry
from feishu_agent_bot.reporting.artifacts import ProfessionalArtifactBuilder
from feishu_agent_bot.reporting.report_ir import ReportIR, ReportSection


FIXTURE_DIR = Path(__file__).with_name("fixtures")
REQUIRED_FIXTURES = {
    "sample_text.pdf",
    "sample_scanned.pdf",
    "sample_prices.csv",
    "sample_competitors.xlsx",
    "sample_metrics.json",
    "sample_report.docx",
    "sample_page.html",
}


def test_alpha_fixed_fixtures_exist_and_cover_all_required_formats():
    _ensure_fixtures()

    existing = {path.name for path in FIXTURE_DIR.iterdir() if path.is_file()}
    assert REQUIRED_FIXTURES <= existing

    detected = {
        name: FileTypeDetector.detect(FIXTURE_DIR / name)
        for name in REQUIRED_FIXTURES
    }
    assert detected == {
        "sample_text.pdf": "pdf",
        "sample_scanned.pdf": "pdf",
        "sample_prices.csv": "csv",
        "sample_competitors.xlsx": "excel",
        "sample_metrics.json": "json",
        "sample_report.docx": "docx",
        "sample_page.html": "html",
    }


def test_alpha_fixed_fixture_e2e_generates_traceable_professional_artifacts(tmp_path):
    _ensure_fixtures()
    working_fixtures = tmp_path / "fixtures"
    working_fixtures.mkdir()
    for name in REQUIRED_FIXTURES:
        shutil.copyfile(FIXTURE_DIR / name, working_fixtures / name)
    registry = default_parser_registry(csv_chunk_size=2)
    parsed_assets = {}
    datasets = []
    profiler = DatasetProfiler()
    data_quality = []
    sources = []

    for index, name in enumerate(sorted(REQUIRED_FIXTURES), start=1):
        path = working_fixtures / name
        file_type = FileTypeDetector.detect(path)
        parsed = registry.parse(file_type, path, asset_id=f"ASSET{index}")
        parsed_assets[name] = parsed
        sources.append(
            {
                "source_id": f"SRC{index}",
                "title": parsed.title or name,
                "canonical_url": f"fixture://{name}",
                "content_type": file_type,
            }
        )
        for table in parsed.tables:
            dataset = TabularDataset(
                dataset_id=f"D{len(datasets) + 1:03d}",
                job_id="alpha-e2e",
                asset_id=parsed.asset_id,
                table_id=table.table_id,
                name=table.caption or table.table_id,
                columns=table.columns,
                rows=table.rows,
                lineage={
                    "source_locator": table.source_locator,
                    "extraction_method": table.extraction_method,
                    "table_metadata": table.metadata,
                },
            )
            profile = profiler.profile(dataset)
            datasets.append(
                {
                    "dataset_id": dataset.dataset_id,
                    "table_id": dataset.table_id,
                    "dataset_name": dataset.name,
                    "columns": dataset.columns,
                    "rows": dataset.rows,
                    "row_count": len(dataset.rows),
                    "column_count": len(dataset.columns),
                    "lineage": dataset.lineage,
                    "profile": profile.__dict__ | {"schema": [item.__dict__ for item in profile.schema]},
                }
            )
            data_quality.append(
                {
                    "dataset_id": dataset.dataset_id,
                    "name": dataset.name,
                    "row_count": len(dataset.rows),
                    "column_count": len(dataset.columns),
                    "profile": profile.__dict__ | {"schema": [item.__dict__ for item in profile.schema]},
                }
            )

    pricing_dataset = _dataset_with_column(datasets, "price_cny")
    competitor_dataset = _dataset_with_column(datasets, "plan / monthly_price_cny")
    json_dataset = _dataset_with_column(datasets, "active_users")
    html_dataset = _dataset_with_column(datasets, "pricing / monthly_cny")
    assert pricing_dataset and competitor_dataset and json_dataset and html_dataset

    analysis_results = [
        {
            "analysis_result_id": "AR_PRICE",
            "skill_name": "pricing_and_packaging",
            "tool_name": "pricing_normalizer",
            "summary": "价格数据来自 CSV、Excel 和 HTML fixture。",
            "input_dataset_ids": [
                pricing_dataset["dataset_id"],
                competitor_dataset["dataset_id"],
                html_dataset["dataset_id"],
            ],
            "input_evidence_ids": ["E_PDF", "E_XLSX"],
            "metrics": {"normalized_price_count": 4, "lowest_price_cny": 89},
            "tables": [
                {
                    "dataset_id": competitor_dataset["dataset_id"],
                    "name": "normalized_pricing",
                    "source_locator": competitor_dataset["lineage"]["source_locator"],
                    "rows": [
                        {
                            "company": "Competitor A",
                            "normalized_price": 109.0,
                            "currency": "CNY",
                            "source_locator": competitor_dataset["lineage"]["source_locator"],
                        },
                        {
                            "company": "Competitor B",
                            "normalized_price": 95.0,
                            "currency": "CNY",
                            "source_locator": html_dataset["lineage"]["source_locator"],
                        },
                    ],
                }
            ],
            "charts": [],
            "limitations": [],
        },
        {
            "analysis_result_id": "AR_TREND",
            "skill_name": "trend_and_change",
            "tool_name": "growth_metrics_calculator",
            "summary": "JSON fixture 提供时间序列增长数据。",
            "input_dataset_ids": [json_dataset["dataset_id"]],
            "input_evidence_ids": ["E_JSON"],
            "metrics": {"latest_active_users": 11200, "latest_growth_pct": 12},
            "tables": [],
            "charts": [
                {
                    "chart_id": "active_users_trend",
                    "chart_type": "line",
                    "title": "活跃用户趋势",
                    "unit": "users",
                    "points": [
                        {"x": "2026-01", "y": 10000},
                        {"x": "2026-02", "y": 11200},
                    ],
                    "dataset_ids": [json_dataset["dataset_id"]],
                    "analysis_result_ids": ["AR_TREND"],
                    "source_ids": ["SRC_JSON"],
                }
            ],
            "limitations": [],
        },
    ]

    ir = ReportIR(
        job_id="alpha-e2e",
        report_id="alpha-e2e-report",
        version=1,
        title="Alpha 固定验收报告",
        sections=[
            ReportSection(
                "S_FEATURES",
                "竞品功能说明",
                "PDF 和 DOCX fixture 说明 Competitor A 支持 fast charging，Competitor B 强调低价。",
                evidence_ids=["E_PDF", "E_DOCX"],
            ),
            ReportSection(
                "S_PRICING",
                "价格与套餐矩阵",
                "CSV、Excel 和 HTML fixture 均指向 109 与 95 这组核心套餐价格。",
                analysis_result_ids=["AR_PRICE"],
                table_ids=[competitor_dataset["table_id"], html_dataset["table_id"]],
                chart_ids=["active_users_trend"],
                evidence_ids=["E_XLSX", "E_HTML"],
            ),
        ],
        tables=[
            {
                "table_id": item["table_id"],
                "dataset_id": item["dataset_id"],
                "title": item["dataset_name"],
                "columns": item["columns"],
                "rows": item["rows"],
                "row_count": item["row_count"],
                "lineage": item["lineage"],
            }
            for item in datasets
        ],
        charts=analysis_results[1]["charts"],
        analysis_results=analysis_results,
        evidence_references=[
            {
                "evidence_id": "E_PDF",
                "source_id": "SRC_PDF",
                "source_locator": "sample_text.pdf#page=1",
                "exact_quote": "Competitor A supports fast charging.",
            },
            {
                "evidence_id": "E_XLSX",
                "source_id": "SRC_XLSX",
                "source_locator": competitor_dataset["lineage"]["source_locator"],
                "exact_quote": "Competitor A Pro monthly price is 109.",
            },
            {
                "evidence_id": "E_JSON",
                "source_id": "SRC_JSON",
                "source_locator": json_dataset["lineage"]["source_locator"],
                "exact_quote": "active_users reaches 11200.",
            },
            {
                "evidence_id": "E_DOCX",
                "source_id": "SRC_DOCX",
                "source_locator": "sample_report.docx#paragraph-2",
                "exact_quote": "中端套餐增长来自企业移动办公场景。",
            },
            {
                "evidence_id": "E_HTML",
                "source_id": "SRC_HTML",
                "source_locator": html_dataset["lineage"]["source_locator"],
                "exact_quote": "官方套餐 monthly_cny 95.",
            },
        ],
        claims=[
            {
                "claim_id": "C1",
                "statement": "竞品套餐价格和功能均可追溯到固定 fixture。",
                "supporting_evidence_ids": ["E_PDF", "E_XLSX", "E_HTML"],
                "contradicting_evidence_ids": [],
            }
        ],
        sources=[
            {"source_id": "SRC_PDF", "title": "PDF 功能说明", "canonical_url": "https://fixture.local/sample_text.pdf"},
            {"source_id": "SRC_XLSX", "title": "Excel 竞品矩阵", "canonical_url": "https://fixture.local/sample_competitors.xlsx"},
            {"source_id": "SRC_JSON", "title": "JSON 时间序列", "canonical_url": "https://fixture.local/sample_metrics.json"},
            {"source_id": "SRC_DOCX", "title": "DOCX 行业报告", "canonical_url": "https://fixture.local/sample_report.docx"},
            {"source_id": "SRC_HTML", "title": "HTML 官方页面", "canonical_url": "https://fixture.local/sample_page.html"},
        ],
        data_quality=data_quality,
        limitations=["固定 fixture 用于无公网验收，不代表真实联网研究。"],
    )

    artifacts = ProfessionalArtifactBuilder(tmp_path).build(ir, ["json", "xlsx", "pdf"])
    by_type = {artifact.artifact_type: artifact for artifact in artifacts}
    job_dir = tmp_path / "alpha-e2e" / "professional" / "v1"

    assert (job_dir / "report_ir.json").is_file()
    assert (job_dir / "report.tex").is_file()
    assert (job_dir / "report.pdf").is_file()
    assert (job_dir / "report.xlsx").is_file()
    assert (job_dir / "charts").is_dir()
    assert (job_dir / "datasets").is_dir()
    assert (job_dir / "manifest.json").is_file()
    assert (job_dir / "artifact_manifest.json").is_file()
    assert by_type["pdf"].status == "ready"

    with fitz.open(job_dir / "report.pdf") as document:
        assert document.page_count >= 1
        pdf_text = "\n".join(page.get_text() for page in document)
        embedded_image_count = sum(
            len(page.get_images(full=True)) for page in document
        )
    assert "Alpha" in pdf_text
    assert "109" in pdf_text
    assert "Competitor A" in pdf_text
    assert embedded_image_count >= 1

    report_tex = (job_dir / "report.tex").read_text(encoding="utf-8")
    assert "\\tableofcontents" in report_tex
    assert "\\begin{longtable}" in report_tex
    assert "charts/active_users_trend.png" in report_tex

    workbook = load_workbook(job_dir / "report.xlsx", read_only=False)
    try:
        assert "Competitor Matrix" in workbook.sheetnames
        assert "Pricing" in workbook.sheetnames
        pricing_rows = list(workbook["Pricing"].iter_rows(values_only=True))
        assert any(109 in row for row in pricing_rows)
        evidence_headers = [cell.value for cell in workbook["Evidence"][1]]
        locator_column = evidence_headers.index("source_locator")
        evidence_rows = list(workbook["Evidence"].iter_rows(min_row=2, values_only=True))
        assert any("sample_text.pdf#page=1" == row[locator_column] for row in evidence_rows)
        assert any("sample_competitors.xlsx#sheet=" in str(row[locator_column]) for row in evidence_rows)
        audit_rows = list(workbook["Analysis Audit"].iter_rows(values_only=True))
        assert any(row[0] == "AR_PRICE" and row[3] == "metrics.lowest_price_cny" and row[4] == 89 for row in audit_rows)
    finally:
        workbook.close()

    chart_metadata = json.loads(
        (job_dir / "charts" / "active_users_trend_metadata.json").read_text(encoding="utf-8")
    )
    assert chart_metadata["dataset_ids"] == [json_dataset["dataset_id"]]
    assert chart_metadata["analysis_result_ids"] == ["AR_TREND"]
    chart_data = (job_dir / "charts" / "active_users_trend_data.csv").read_text(encoding="utf-8")
    assert "2026-02,11200" in chart_data

    dataset_files = sorted(path.name for path in (job_dir / "datasets").iterdir())
    assert any(name.endswith(".csv") for name in dataset_files)
    manifest = json.loads((job_dir / "manifest.json").read_text(encoding="utf-8"))
    artifact_paths = {item["path"] for item in manifest["artifacts"]}
    assert "report_ir.json" in artifact_paths
    assert "report.xlsx" in artifact_paths
    assert "report.pdf" in artifact_paths
    assert "charts/active_users_trend_data.csv" in artifact_paths
    assert any(path.startswith("datasets/") for path in artifact_paths)

    before = sorted(path.relative_to(tmp_path).as_posix() for path in tmp_path.rglob("*"))
    second = ProfessionalArtifactBuilder(tmp_path).build(ir, ["json", "xlsx", "pdf"])
    after = sorted(path.relative_to(tmp_path).as_posix() for path in tmp_path.rglob("*"))
    assert before == after
    assert {artifact.artifact_type for artifact in second} == {artifact.artifact_type for artifact in artifacts}
    assert [
        (artifact.artifact_type, artifact.content_hash, artifact.status)
        for artifact in second
    ] == [
        (artifact.artifact_type, artifact.content_hash, artifact.status)
        for artifact in artifacts
    ]
    assert not (tmp_path / "alpha-e2e" / "professional" / "v2").exists()


def _ensure_fixtures() -> None:
    missing = [name for name in REQUIRED_FIXTURES if not (FIXTURE_DIR / name).is_file()]
    if missing:
        subprocess.run(
            [".venv/bin/python", "scripts/build_e2e_fixtures.py"],
            cwd=Path(__file__).resolve().parents[1],
            check=True,
        )


def _dataset_with_column(datasets: list[dict], column: str) -> dict | None:
    for dataset in datasets:
        if column in dataset["columns"]:
            return dataset
    return None
